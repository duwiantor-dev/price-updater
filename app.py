import io
import re
import zipfile
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple

import streamlit as st
from openpyxl import load_workbook, Workbook
from openpyxl.cell.cell import MergedCell


# =========================
# Config / Constants
# =========================
SKU_SPLIT_PLUS = re.compile(r"\+")

SMALL_TO_THOUSAND_THRESHOLD = 1_000_000
AUTO_MULTIPLIER_FOR_SMALL = 1000

# Pricelist: header minimal yang kita cari
PL_HEADER_SKU_CANDIDATES = ["KODEBARANG", "KODE BARANG", "SKU", "SKU NO", "SKU_NO", "KODEBARANG "]
PRICELIST_HEADER_ROW_FIXED = 2
PL_PRICE_COL_TIKTOK = "M3"
PL_PRICE_COL_SHOPEE = "M4"

# Addon mapping: header yang diharapkan (boleh lebih dari 1 kandidat)
ADDON_CODE_CANDIDATES = ["addon_code", "ADDON_CODE", "Addon Code", "Kode", "KODE", "KODE ADDON", "KODE_ADDON"]
ADDON_PRICE_CANDIDATES = ["harga", "HARGA", "Price", "PRICE", "Harga"]


# =========================
# Utils
# =========================
def normalize_text(x) -> str:
    return "" if x is None else str(x).strip()


def normalize_addon_code(x) -> str:
    return normalize_text(x).upper()


def parse_platform_sku(full_sku: str) -> Tuple[str, List[str]]:
    """SKU bisa 'BASE+ADDON1+ADDON2'"""
    if full_sku is None:
        return "", []
    s = str(full_sku).strip()
    if not s:
        return "", []
    parts = SKU_SPLIT_PLUS.split(s)
    base = parts[0].strip()
    addons = [p.strip() for p in parts[1:] if p and p.strip()]
    return base, addons


def parse_number_like_id(x) -> str:
    """Biar ID yang kebaca float di Excel (mis 1.0) jadi '1'."""
    if x is None:
        return ""
    if isinstance(x, int):
        return str(x)
    if isinstance(x, float):
        if x.is_integer():
            return str(int(x))
        return str(x)
    return str(x).strip()


def parse_price_cell(val) -> Optional[int]:
    if val is None:
        return None

    if isinstance(val, (int, float)):
        try:
            if isinstance(val, float) and val.is_integer():
                return int(val)
            return int(round(float(val)))
        except Exception:
            return None

    s = str(val).strip()
    if not s:
        return None

    s = s.replace("Rp", "").replace("rp", "").replace(" ", "")

    # Handle format 1.234.567,00 / 1.234.567 / 1,234,567
    if "." in s and "," in s:
        s = s.replace(".", "").replace(",", ".")
    elif "." in s and "," not in s:
        s = s.replace(".", "")
    elif "," in s and "." not in s:
        s = s.replace(",", "")

    try:
        f = float(s)
        if f.is_integer():
            return int(f)
        return int(round(f))
    except Exception:
        return None


def apply_multiplier_if_needed(x: int) -> int:
    if x is None:
        return 0
    if x < SMALL_TO_THOUSAND_THRESHOLD:
        return x * AUTO_MULTIPLIER_FOR_SMALL
    return x


def safe_set_cell_value(ws, row: int, col: int, value):
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        coord = cell.coordinate
        for merged in ws.merged_cells.ranges:
            if coord in merged:
                ws.cell(row=merged.min_row, column=merged.min_col).value = value
                return
        return
    cell.value = value


def workbook_to_bytes(wb) -> bytes:
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# =========================
# Dynamic header scan (lebih robust)
# =========================
def _norm(x) -> str:
    """Normalize header cell text: tahan newline, spasi ganda, dll."""
    if x is None:
        return ""
    s = str(x)
    s = s.replace("\n", " ")
    s = re.sub(r"\s+", " ", s)  # collapse spasi/tab/newline
    return s.strip().lower()


def _match_col(lower_to_col: Dict[str, int], candidates: List[str]) -> Optional[int]:
    """Cari kolom berdasarkan kandidat header:
    1) exact match
    2) contains match (header ada tambahan kata seperti '(Optional)')
    """
    if not candidates:
        return None

    # exact
    for cand in candidates:
        c = _norm(cand)
        if c in lower_to_col:
            return lower_to_col[c]

    # contains
    keys = list(lower_to_col.keys())
    for cand in candidates:
        c = _norm(cand)
        if not c:
            continue
        for k in keys:
            if c in k or k in c:
                return lower_to_col[k]

    return None


def find_header_cols_and_data_start(
    ws,
    sku_candidates: List[str],
    price_candidates: List[str],
    scan_rows: int = 120,
) -> Tuple[int, int, int, int]:
    """
    Return: (header_row, sku_col, price_col, data_start_row)
    """
    max_r = min(scan_rows, ws.max_row)

    for r in range(1, max_r + 1):
        lower_to_col: Dict[str, int] = {}
        for c in range(1, ws.max_column + 1):
            v = _norm(ws.cell(row=r, column=c).value)
            if v and v not in lower_to_col:
                lower_to_col[v] = c

        sku_col = _match_col(lower_to_col, sku_candidates)
        price_col = _match_col(lower_to_col, price_candidates)

        if sku_col and price_col:
            # detect first data row
            data_start = None
            for rr in range(r + 1, ws.max_row + 1):
                sku_val = ws.cell(row=rr, column=sku_col).value
                if parse_number_like_id(sku_val):
                    data_start = rr
                    break
            if data_start is None:
                data_start = r + 1
            return r, sku_col, price_col, data_start

    raise ValueError("Header SKU/Price tidak ketemu (dynamic scan gagal).")


def find_col_on_header_row(ws, header_row: int, candidates: Optional[List[str]]) -> Optional[int]:
    if not candidates:
        return None

    lower_to_col: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = _norm(ws.cell(row=header_row, column=c).value)
        if v and v not in lower_to_col:
            lower_to_col[v] = c

    return _match_col(lower_to_col, candidates)


# =========================
# Pricelist & Addon
# =========================
def find_header_row_and_cols_pricelist(ws) -> Tuple[int, int, int, int]:
    r = PRICELIST_HEADER_ROW_FIXED
    candidates = [_norm(c) for c in PL_HEADER_SKU_CANDIDATES]
    target_m3 = _norm(PL_PRICE_COL_TIKTOK)
    target_m4 = _norm(PL_PRICE_COL_SHOPEE)

    lower_to_col: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = _norm(ws.cell(row=r, column=c).value)
        if v and v not in lower_to_col:
            lower_to_col[v] = c

    sku_col = None
    for cand in candidates:
        if cand in lower_to_col:
            sku_col = lower_to_col[cand]
            break

    if sku_col is None or target_m3 not in lower_to_col or target_m4 not in lower_to_col:
        raise ValueError(
            f"Header Pricelist row {PRICELIST_HEADER_ROW_FIXED} tidak sesuai. "
            f"Pastikan ada kolom KODEBARANG (atau setara) dan kolom M3 & M4."
        )

    return r, sku_col, lower_to_col[target_m3], lower_to_col[target_m4]


def load_pricelist_map(pl_bytes: bytes) -> Dict[str, Dict[str, int]]:
    wb = load_workbook(io.BytesIO(pl_bytes), data_only=True)
    ws = wb.active

    header_row, sku_col, m3_col, m4_col = find_header_row_and_cols_pricelist(ws)

    m: Dict[str, Dict[str, int]] = {}
    for r in range(header_row + 1, ws.max_row + 1):
        sku = normalize_text(ws.cell(row=r, column=sku_col).value)
        if not sku:
            continue

        m3_raw = parse_price_cell(ws.cell(row=r, column=m3_col).value)
        m4_raw = parse_price_cell(ws.cell(row=r, column=m4_col).value)

        if m3_raw is None and m4_raw is None:
            continue

        m[sku] = {}
        if m3_raw is not None:
            m[sku]["M3"] = int(apply_multiplier_if_needed(m3_raw))
        if m4_raw is not None:
            m[sku]["M4"] = int(apply_multiplier_if_needed(m4_raw))

    return m


def load_addon_map(addon_bytes: bytes) -> Dict[str, int]:
    wb = load_workbook(io.BytesIO(addon_bytes), data_only=True)
    ws = wb.active

    header_row = None
    code_col = None
    price_col = None

    code_cands = [_norm(c) for c in ADDON_CODE_CANDIDATES]
    price_cands = [_norm(c) for c in ADDON_PRICE_CANDIDATES]

    for r in range(1, 50):
        lower_to_col: Dict[str, int] = {}
        for c in range(1, ws.max_column + 1):
            v = _norm(ws.cell(row=r, column=c).value)
            if v and v not in lower_to_col:
                lower_to_col[v] = c

        found_code = next((lower_to_col.get(cc) for cc in code_cands if cc in lower_to_col), None)
        found_price = next((lower_to_col.get(pc) for pc in price_cands if pc in lower_to_col), None)

        if found_code and found_price:
            header_row, code_col, price_col = r, found_code, found_price
            break

    if header_row is None or code_col is None or price_col is None:
        raise ValueError("Header Addon Mapping tidak ketemu. Pastikan ada kolom addon_code & harga (atau setara).")

    m: Dict[str, int] = {}
    for r in range(header_row + 1, ws.max_row + 1):
        code = normalize_addon_code(ws.cell(row=r, column=code_col).value)
        if not code:
            continue

        price_raw = parse_price_cell(ws.cell(row=r, column=price_col).value)
        if price_raw is None:
            continue

        m[code] = int(apply_multiplier_if_needed(int(price_raw)))

    return m


# =========================
# Core compute
# =========================
@dataclass
class RowChange:
    file: str
    excel_row: int
    sku_full: str
    old_price: int
    new_price: int
    reason: str


def compute_new_price_for_row(
    sku_full: str,
    pl_map: Dict[str, Dict[str, int]],
    addon_map: Dict[str, int],
    discount_rp: int,
    price_key: str,
) -> Tuple[Optional[int], str]:
    base_sku, addons = parse_platform_sku(sku_full)
    if not base_sku:
        return None, "SKU kosong"

    pl = pl_map.get(base_sku)
    if not pl:
        return None, "Base SKU tidak ada di Pricelist"

    base_price = pl.get(price_key)
    if base_price is None:
        return None, f"Harga {price_key} kosong di Pricelist"

    addon_total = 0
    for a in addons:
        code = normalize_addon_code(a)
        if not code:
            continue
        if code not in addon_map:
            return None, f"Addon '{code}' tidak ada di file Addon Mapping"
        addon_total += int(addon_map[code])

    final_price = int(base_price) + int(addon_total) - int(discount_rp)
    if final_price < 0:
        final_price = 0

    return final_price, f"{price_key} + addon - diskon"


def keep_only_changed_rows_in_place(ws, data_start_row: int, changed_row_numbers: List[int]):
    keep = set(changed_row_numbers)
    for r in range(ws.max_row, data_start_row - 1, -1):
        if r not in keep:
            ws.delete_rows(r, 1)


def make_report_workbook(changes: List[RowChange]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "changes_report"
    ws.append(["file", "row", "sku_full", "old_price", "new_price", "reason"])
    for ch in changes:
        ws.append([ch.file, ch.excel_row, ch.sku_full, ch.old_price, ch.new_price, ch.reason])
    return workbook_to_bytes(wb)


# =========================
# Template Profiles (sesuai penjelasan kamu)
# =========================
@dataclass
class TemplateProfile:
    key: str
    label: str
    mode_output: str            # "inplace" | "generate"
    default_price_key: str      # "M3" | "M4"

    # input headers
    sku_headers: List[str]
    price_headers: List[str]

    # generate only: ambil ID produk & ID SKU dari header input
    product_id_headers: Optional[List[str]] = None
    sku_id_headers: Optional[List[str]] = None

    # output headers for generate
    out_product_header: Optional[str] = None
    out_sku_header: Optional[str] = None
    out_price_header: Optional[str] = None


# INPUT HEADER sesuai requirement kamu:
# - TikTok normal & PM normal: "SKU Penjual", "harga ritel (mata uang lokal)"
HDR_TT_PM_SKU = ["SKU Penjual", "Sku Penjual", "Seller SKU"]
HDR_TT_PM_PRICE = [
    "Harga Ritel (Mata Uang Lokal)",
    "harga ritel (mata uang lokal)",
    "Harga Ritel",
    "harga ritel",
]

# - Shopee normal: "SKU", "harga"
HDR_SHOPEE_NORMAL_SKU = ["SKU", "Sku"]
HDR_SHOPEE_NORMAL_PRICE = ["harga", "Harga", "price", "Price"]

# Shopee diskon (template sering beda; dibuat lebih gemuk)
HDR_SHOPEE_DISKON_SKU = [
    "SKU",
    "SKU Penjual",
    "SKU Ref. No.(Optional)",
    "SKU Ref. No. (Optional)",
    "SKU Ref No",
    "Seller SKU",
]
HDR_SHOPEE_DISKON_PRICE = [
    "harga",
    "Harga",
    "Harga diskon",
    "Harga Diskon",
    "Discount Price",
    "Harga Ritel (Mata Uang Lokal)",
]

# ID headers (untuk output 3 kolom)
HDR_PRODUCT_ID = [
    "ID produk",
    "ID Produk",
    "ID Produk (wajib)",
    "produk_id",
    "Product_id",
    "Product ID",
    "ID Product",
]
HDR_SKU_ID = [
    "ID SKU",
    "ID SKU (wajib)",
    "SKU_id",
    "SKU ID",
    "ID varian",
    "ID Varian",
    "Variant ID",
    "ID Variasi",
]

TEMPLATES: List[TemplateProfile] = [
    TemplateProfile(
        key="NORMAL_TIKTOK",
        label="Normal TikTok (pakai M3)",
        mode_output="inplace",
        default_price_key="M3",
        sku_headers=HDR_TT_PM_SKU,
        price_headers=HDR_TT_PM_PRICE,
    ),
    TemplateProfile(
        key="NORMAL_SHOPEE",
        label="Normal Shopee (pakai M4)",
        mode_output="inplace",
        default_price_key="M4",
        sku_headers=HDR_SHOPEE_NORMAL_SKU,
        price_headers=HDR_SHOPEE_NORMAL_PRICE,
    ),
    TemplateProfile(
        key="NORMAL_PM",
        label="Normal PM (pakai M4)",
        mode_output="inplace",
        default_price_key="M4",
        sku_headers=HDR_TT_PM_SKU,
        price_headers=HDR_TT_PM_PRICE,
    ),

    # Diskon/Coret TikTok: output 3 kolom wajib
    TemplateProfile(
        key="CORET_TIKTOK",
        label="Diskon / Harga Coret TikTok (pakai M3) - output beda",
        mode_output="generate",
        default_price_key="M3",
        sku_headers=HDR_TT_PM_SKU,
        price_headers=HDR_TT_PM_PRICE,
        product_id_headers=HDR_PRODUCT_ID,
        sku_id_headers=HDR_SKU_ID,
        out_product_header="produk_id (wajib)",
        out_sku_header="SKU_id (wajib)",
        out_price_header="Harga Penawaran (wajib)",
    ),

    # Diskon/Coret Shopee: sesuai penjelasan kamu -> output sama (SKU Penjual + harga ritel)
    # Jadi kita proses "inplace" saja.
    TemplateProfile(
        key="CORET_SHOPEE",
        label="Diskon / Harga Coret Shopee (pakai M4) - output beda",
        mode_output="inplace",
        default_price_key="M4",
        sku_headers=HDR_SHOPEE_DISKON_SKU,
        price_headers=HDR_SHOPEE_DISKON_PRICE,
    ),

    # Diskon/Coret PM: output 3 kolom wajib
    TemplateProfile(
        key="CORET_PM",
        label="Diskon / Harga Coret PM (pakai M4) - output beda",
        mode_output="generate",
        default_price_key="M4",
        sku_headers=HDR_TT_PM_SKU,
        price_headers=HDR_TT_PM_PRICE,
        product_id_headers=HDR_PRODUCT_ID,
        sku_id_headers=HDR_SKU_ID,
        out_product_header="produk_id (wajib)",
        out_sku_header="SKU_id (wajib)",
        out_price_header="Harga Penawaran (wajib)",
    ),
]


def build_output_workbook_3cols(tpl: TemplateProfile, rows: List[Dict[str, object]]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws.append([tpl.out_product_header, tpl.out_sku_header, tpl.out_price_header])

    for r in rows:
        ws.append([
            r.get("product_id", ""),
            r.get("sku_id", ""),
            r.get("price", ""),
        ])
    return wb


# =========================
# UI
# =========================
st.set_page_config(page_title="Harga (Dynamic Template)", layout="wide")
st.title("Harga (Dynamic Template)")

tpl_map = {t.label: t for t in TEMPLATES}
tpl_label = st.selectbox("Pilih Template Proses", options=list(tpl_map.keys()))
tpl = tpl_map[tpl_label]

c1, c2, c3 = st.columns(3)
with c1:
    mass_files = st.file_uploader("Upload file input (bisa banyak)", type=["xlsx"], accept_multiple_files=True)
with c2:
    pl_file = st.file_uploader("Upload Pricelist (fixed)", type=["xlsx"])
with c3:
    addon_file = st.file_uploader("Upload Addon Mapping (fixed)", type=["xlsx"])

st.divider()

colA, colB = st.columns([1, 2])
with colA:
    discount_rp = st.number_input("Diskon (Rp)", min_value=0, value=0, step=1000)

with colB:
    override = st.checkbox("Override price_key manual (opsional)")
    if override:
        price_key_ui = st.radio("Pakai harga Pricelist", options=["M3", "M4"], horizontal=True)
    else:
        price_key_ui = tpl.default_price_key
        st.caption(f"price_key otomatis dari template: **{price_key_ui}**")

process = st.button("Proses")

if process:
    if not mass_files or pl_file is None or addon_file is None:
        st.error("Wajib upload: file input (minimal 1), Pricelist, dan Addon Mapping.")
        st.stop()

    try:
        pl_map = load_pricelist_map(pl_file.getvalue())
    except Exception as e:
        st.error(f"Gagal baca Pricelist: {e}")
        st.stop()

    try:
        addon_map = load_addon_map(addon_file.getvalue())
    except Exception as e:
        st.error(f"Gagal baca Addon Mapping: {e}")
        st.stop()

    changed_rows: List[RowChange] = []
    output_files: List[Tuple[str, bytes]] = []

    for mf in mass_files:
        filename = mf.name
        wb_in = load_workbook(io.BytesIO(mf.getvalue()))
        ws_in = wb_in.active

        # detect header + data start (dynamic robust)
        try:
            header_row, sku_col, price_col, data_start_row = find_header_cols_and_data_start(
                ws_in,
                sku_candidates=tpl.sku_headers,
                price_candidates=tpl.price_headers,
                scan_rows=160,
            )
        except Exception as e:
            changed_rows.append(RowChange(
                file=filename, excel_row=0, sku_full="", old_price=0, new_price=0,
                reason=f"Gagal detect header/data start: {e}",
            ))
            continue

        # optional cols (for generate mode)
        product_col = find_col_on_header_row(ws_in, header_row, tpl.product_id_headers)
        sku_id_col = find_col_on_header_row(ws_in, header_row, tpl.sku_id_headers)

        # ============ MODE: INPLACE ============
        if tpl.mode_output == "inplace":
            changed_row_numbers: List[int] = []

            for r in range(data_start_row, ws_in.max_row + 1):
                sku_full = parse_number_like_id(ws_in.cell(row=r, column=sku_col).value)
                if not sku_full:
                    continue

                old_price_raw = parse_price_cell(ws_in.cell(row=r, column=price_col).value)
                old_price = int(old_price_raw) if old_price_raw is not None else 0

                new_price, reason = compute_new_price_for_row(
                    sku_full=sku_full,
                    pl_map=pl_map,
                    addon_map=addon_map,
                    discount_rp=int(discount_rp),
                    price_key=price_key_ui,
                )

                if new_price is None or int(new_price) == int(old_price):
                    continue

                safe_set_cell_value(ws_in, row=r, col=price_col, value=int(new_price))
                changed_row_numbers.append(r)

                changed_rows.append(RowChange(
                    file=filename, excel_row=r, sku_full=sku_full,
                    old_price=int(old_price), new_price=int(new_price),
                    reason=f"[{tpl.key}] {reason}",
                ))

            if changed_row_numbers:
                keep_only_changed_rows_in_place(ws_in, data_start_row, changed_row_numbers)
                out_bytes = workbook_to_bytes(wb_in)
                out_name = filename.replace(".xlsx", f"_changed_only_{tpl.key}_{price_key_ui}.xlsx")
                output_files.append((out_name, out_bytes))

        # ============ MODE: GENERATE (3 kolom wajib) ============
        else:
            rows_for_output: List[Dict[str, object]] = []

            # validasi kolom wajib dari header
            if product_col is None:
                changed_rows.append(RowChange(
                    file=filename, excel_row=0, sku_full="", old_price=0, new_price=0,
                    reason=f"[{tpl.key}] Kolom produk_id tidak ketemu. Kandidat header: {tpl.product_id_headers}",
                ))
                continue
            if sku_id_col is None:
                changed_rows.append(RowChange(
                    file=filename, excel_row=0, sku_full="", old_price=0, new_price=0,
                    reason=f"[{tpl.key}] Kolom SKU_id tidak ketemu. Kandidat header: {tpl.sku_id_headers}",
                ))
                continue

            for r in range(data_start_row, ws_in.max_row + 1):
                sku_full = parse_number_like_id(ws_in.cell(row=r, column=sku_col).value)
                if not sku_full:
                    continue

                old_price_raw = parse_price_cell(ws_in.cell(row=r, column=price_col).value)
                old_price = int(old_price_raw) if old_price_raw is not None else 0

                new_price, reason = compute_new_price_for_row(
                    sku_full=sku_full,
                    pl_map=pl_map,
                    addon_map=addon_map,
                    discount_rp=int(discount_rp),
                    price_key=price_key_ui,
                )
                if new_price is None or int(new_price) == int(old_price):
                    continue

                product_id = parse_number_like_id(ws_in.cell(row=r, column=product_col).value)
                sku_id = parse_number_like_id(ws_in.cell(row=r, column=sku_id_col).value)

                rows_for_output.append({
                    "product_id": product_id,
                    "sku_id": sku_id,
                    "price": int(new_price),
                })

                changed_rows.append(RowChange(
                    file=filename, excel_row=r, sku_full=sku_full,
                    old_price=int(old_price), new_price=int(new_price),
                    reason=f"[{tpl.key}] {reason}",
                ))

            if rows_for_output:
                wb_out = build_output_workbook_3cols(tpl, rows_for_output)
                out_bytes = workbook_to_bytes(wb_out)
                out_name = filename.replace(".xlsx", f"_OUTPUT_{tpl.key}_{price_key_ui}.xlsx")
                output_files.append((out_name, out_bytes))

    # ===== UI output =====
    st.subheader("Preview (yang berubah saja)")
    if not changed_rows:
        st.info("Tidak ada perubahan harga.")
    else:
        import pandas as pd
        df_preview = pd.DataFrame([{
            "file": x.file,
            "row": x.excel_row,
            "sku_full": x.sku_full,
            "old_price": x.old_price,
            "new_price": x.new_price,
            "reason": x.reason,
        } for x in changed_rows])
        st.dataframe(df_preview.head(500), use_container_width=True)

    st.divider()

    if len(output_files) == 0:
        st.warning("Tidak ada file yang berubah, jadi tidak ada file output untuk didownload.")
    elif len(output_files) == 1:
        name, data = output_files[0]
        st.download_button(
            "Download hasil (XLSX) - hanya yang berubah",
            data=data,
            file_name=name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        zbuf = io.BytesIO()
        with zipfile.ZipFile(zbuf, "w", zipfile.ZIP_DEFLATED) as zf:
            for name, data in output_files:
                zf.writestr(name, data)
            zf.writestr("changes_report.xlsx", make_report_workbook(changed_rows))

        st.download_button(
            "Download semua hasil (ZIP) - hanya yang berubah",
            data=zbuf.getvalue(),
            file_name=f"mass_update_results_{tpl.key}_{price_key_ui}.zip",
            mime="application/zip",
        )

    if changed_rows:
        st.download_button(
            "Download Report Perubahan (XLSX)",
            data=make_report_workbook(changed_rows),
            file_name="changes_report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
